import openpyxl

workbook = openpyxl.Workbook()
workbook.create_sheet()
sheet = workbook.active
sheet['A1']='Приветствую'

print(workbook.sheetnames)

workbook.save('sheets.xlsx')
