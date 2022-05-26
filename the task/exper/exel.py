from openpyxl import load_workbook

wb = load_workbook(filename='exlist.xlsx')
sheet = wb['Лист1']
rows = sheet.max_row
cols = sheet.max_column
clients = []
clients_visited=[]
per = 0
ban = 'мт'

for i in range(1, rows + 1):

    for j in range(1, cols + 1):
        cell = sheet.cell(row=i, column=j)
        if cell.value == None or ban in str(cell.value):
            None
        else:
            if cell.value.lower() in str(clients):

                index = clients.index(cell.value.lower())
                clients_visited[index]=int(clients_visited[index])+1



            else:
                clients.append(cell.value.lower())
                clients_visited.append(1)

                per += 1
for i in range(len(clients)):
    print(clients[i],    clients_visited[i])
